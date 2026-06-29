import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, time
import openpyxl
import io
import random
import time as time_module

def str_to_timedelta(time_str):
    time_str = str(time_str).strip()
    
    # Quét sạch dấu phẩy/chấm của phần Frame (VD: ,00 hoặc .00) chuyển thành dấu hai chấm
    time_str = time_str.replace(',', ':').replace('.', ':')
    parts = time_str.split(':')
    
    if len(parts) >= 3:
        try:
            return timedelta(hours=int(parts[0]), minutes=int(parts[1]), seconds=int(parts[2]))
        except:
            pass
    elif len(parts) == 2:
        try:
            return timedelta(minutes=int(parts[0]), seconds=int(parts[1]))
        except:
            pass
            
    try:
        t = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
        return timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
    except:
        pass
        
    return timedelta(seconds=0)

def timedelta_to_str(td):
    total_seconds = int(td.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def find_sequence(pool, target_seconds, max_sai_so, max_lap, bypass_short_limit, bypass_error_limit):
    valid_pool = [f for f in pool if f['duration_secs'] > 0]
    if not valid_pool:
        return None, "❌ Kho dữ liệu rỗng! Không tìm thấy file hợp lệ nào.", 0
        
    best_path = None
    best_error = 999999
    best_has_priority = False
    
    start_time = time_module.time()
    
    while time_module.time() - start_time < 2.5:
        force_short_count = random.choice([0, 1, 2])
        curr_sum = 0
        short_sum = 0
        path = []
        usage = {}
        
        shuffled_pool = list(valid_pool)
        random.shuffle(shuffled_pool)
        
        while True:
            candidates = []
            for f in shuffled_pool:
                if path and f['name'] == path[-1]['name']:
                    continue
                dur = f['duration_secs']
                is_short = dur <= 60
                
                # File > 5 phút (300s) tuyệt đối không lặp lại
                if is_short:
                    limit = max_lap
                elif dur > 300:
                    limit = 1
                else:
                    limit = 2
                    
                if usage.get(f['name'], 0) >= limit:
                    continue
                    
                if not bypass_short_limit and is_short and short_sum + dur >= 900:
                    continue
                candidates.append(f)
                
            if not candidates:
                break
                
            chosen = None
            if len(path) < force_short_count:
                shorts = [c for c in candidates if c['duration_secs'] <= 60]
                if shorts:
                    chosen = random.choice(shorts)
                    
            if not chosen:
                fits_well = [c for c in candidates if curr_sum + c['duration_secs'] <= target_seconds + max_sai_so]
                
                if fits_well:
                    fits_well.sort(key=lambda x: x['duration_secs'], reverse=True)
                    top_k = min(3, len(fits_well))
                    chosen = random.choice(fits_well[:top_k])
                else:
                    if bypass_error_limit and candidates:
                        candidates.sort(key=lambda x: x['duration_secs'])
                        chosen = candidates[0]
                    else:
                        break 
                    
            path.append(chosen)
            curr_sum += chosen['duration_secs']
            if chosen['duration_secs'] <= 60:
                short_sum += chosen['duration_secs']
            usage[chosen['name']] = usage.get(chosen['name'], 0) + 1
            
            error = curr_sum - target_seconds
            is_acceptable_error = bypass_error_limit or (-max_sai_so <= error <= max_sai_so)
            
            if is_acceptable_error:
                abs_err = abs(error)
                start_short = 0
                for p in path:
                    if p['duration_secs'] <= 60:
                        start_short += 1
                    else:
                        break
                is_priority = 1 <= start_short <= 2
                
                if abs_err < best_error:
                    best_error = abs_err
                    best_path = list(path)
                    best_has_priority = is_priority
                elif abs_err == best_error:
                    if is_priority and not best_has_priority:
                        best_path = list(path)
                        best_has_priority = True
                        
            if curr_sum >= target_seconds:
                break

    if best_path is not None:
        actual_sum = sum(f['duration_secs'] for f in best_path)
        sai_so = actual_sum - target_seconds
        return best_path, "Thành công", sai_so
    else:
        err_msg = (f"❌ ỨNG DỤNG TỪ CHỐI TẠO FILE do các điều kiện cài đặt không thể thỏa mãn.\n\n"
                   f"👉 **MẸO XỬ LÝ NHANH:** Vui lòng tích chọn **'Bỏ qua giới hạn sai số'** để ép hệ thống tự động xuất file bằng mọi giá!")
        return None, err_msg, 0


# ==== GIAO DIỆN STREAMLIT ====
st.set_page_config(page_title="App Lịch Chèn Sự Cố", layout="wide")
st.title("Ứng dụng Lịch Chèn Sự Cố")
st.header("CÁC DỮ LIỆU NHẬP VÀO")

col1, col2, col3, col4 = st.columns(4)
with col1: gio_input_cg = st.text_input("1. Giờ CG Phim hiện tại", "12:00:00")
with col2: gio_mat_tin_hieu = st.text_input("2. Giờ mất tín hiệu", "13:30:00")
with col3: gio_input_next = st.text_input("3. Giờ CG phim tiếp theo", "14:00:00")

td_cg = str_to_timedelta(gio_input_cg)
td_mat = str_to_timedelta(gio_mat_tin_hieu)
td_next = str_to_timedelta(gio_input_next)

tab1, tab2 = st.tabs(["I. MÀN INPUT", "II. MÀN OUTPUT"])

with tab1:
    diff_mat_cg = td_mat - td_cg
    
    if td_mat >= td_next:
        st.error("⚠️ **CẢNH BÁO:** Hãy nhập giờ CG phim tiếp theo 2 > Giờ mất tín hiệu.")
        
        with col4:
            gio_input_next_2 = st.text_input("3b. Giờ CG phim tiếp theo 2", "15:00:00")
        td_next_2 = str_to_timedelta(gio_input_next_2)
        
        if diff_mat_cg >= timedelta(hours=1):
            gio_chen = td_mat
            thoi_luong = td_next_2 - td_mat if td_next_2 > td_mat else timedelta(seconds=0)
        else:
            gio_chen = td_next
            thoi_luong = td_next_2 - td_next if td_next_2 > td_next else timedelta(seconds=0)
    else:
        with col4:
            st.write("") 
            
        if diff_mat_cg >= timedelta(hours=1):
            gio_chen = td_mat
            thoi_luong = td_next - td_mat
        else:
            gio_chen = td_cg
            thoi_luong = td_next - td_cg
            
    st.success(f"4. GIỜ CHÈN SỰ CỐ: **{timedelta_to_str(gio_chen)}**")
    st.success(f"5. THỜI LƯỢNG CHÈN: **{timedelta_to_str(thoi_luong)}**")
    
    if thoi_luong.total_seconds() > 0:
        if st.button("Áp dụng dữ liệu từ Màn INPUT cho File"):
            st.session_state['active_gio_chen'] = gio_chen
            st.session_state['active_thoi_luong'] = thoi_luong
            st.success("Đã chọn dữ liệu Màn INPUT!")
    else:
        st.warning("Thời lượng chèn bằng 0, vui lòng kiểm tra hoặc điều chỉnh lại giờ tiếp theo cho hợp lý.")

with tab2:
    gio_chen_out = td_mat + timedelta(minutes=5) - timedelta(hours=1)
    thoi_luong_out = td_next - gio_chen_out
    st.success(f"4. GIỜ CHÈN SỰ CỐ: **{timedelta_to_str(gio_chen_out)}**")
    st.success(f"5. THỜI LƯỢNG CHÈN: **{timedelta_to_str(thoi_luong_out)}**")
    if st.button("Áp dụng dữ liệu từ Màn OUTPUT cho File"):
        st.session_state['active_gio_chen'] = gio_chen_out
        st.session_state['active_thoi_luong'] = thoi_luong_out
        st.success("Đã chọn dữ liệu Màn OUTPUT!")

st.divider()
st.header("6. CẤU HÌNH AI & XUẤT FILE")

if 'active_gio_chen' not in st.session_state:
    st.warning("Vui lòng bấm 'Áp dụng dữ liệu' ở Màn INPUT hoặc OUTPUT trước khi xuất file.")
else:
    active_gio = st.session_state['active_gio_chen']
    active_tl = st.session_state['active_thoi_luong']
    
    st.write(f"**Giờ chèn đang chọn:** {timedelta_to_str(active_gio)} | **Thời lượng yêu cầu:** {timedelta_to_str(active_tl)}")
    
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        user_sai_so = st.slider("Sai số tối đa thông thường (giây)", min_value=0, max_value=120, value=20, step=5)
    with col_c2:
        user_max_lap = st.slider("Số lần lặp tối đa của 1 file ngắn", min_value=2, max_value=10, value=3, step=1)
        
    with st.expander("🛠️ TÙY CHỌN BỎ QUA ĐIỀU KIỆN", expanded=True):
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            bypass_short = st.checkbox("Bỏ qua giới hạn 15 phút cho file ngắn", value=False)
        with col_b2:
            bypass_error = st.checkbox("Bỏ qua giới hạn sai số", value=False)
    
    uploaded_file = st.file_uploader("Tải lên file Excel mẫu (.xlsx)", type=["xlsx"])
    if uploaded_file is not None:
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        file_pool = []
        for row in range(7, sheet.max_row + 1):
            file_name = sheet.cell(row=row, column=3).value
            file_dur = sheet.cell(row=row, column=4).value
            file_type = sheet.cell(row=row, column=5).value
            
            if file_name and file_dur is not None:
                try:
                    if isinstance(file_dur, str):
                        dur_td = str_to_timedelta(file_dur)
                    elif isinstance(file_dur, timedelta):
                        dur_td = file_dur
                    elif isinstance(file_dur, time):
                        dur_td = timedelta(hours=file_dur.hour, minutes=file_dur.minute, seconds=file_dur.second)
                    elif isinstance(file_dur, datetime):
                        dur_td = timedelta(hours=file_dur.hour, minutes=file_dur.minute, seconds=file_dur.second)
                    elif isinstance(file_dur, (int, float)): 
                        total_seconds = int(file_dur * 86400)
                        dur_td = timedelta(seconds=total_seconds)
                    else:
                        dur_td = str_to_timedelta(str(file_dur))
                except Exception:
                    continue
                    
                sec = int(dur_td.total_seconds())
                if sec > 0:
                    file_pool.append({
                        'name': file_name,
                        'duration_str': timedelta_to_str(dur_td),
                        'duration_secs': sec,
                        'type': file_type if file_type else "CM"
                    })
        
        if st.button("7. XUẤT FILE CHÈN", type="primary"):
            target_secs = int(active_tl.total_seconds())
            with st.spinner("AI đang tính toán tổ hợp lịch chèn..."):
                sequence, msg, sai_so = find_sequence(file_pool, target_secs, user_sai_so, user_max_lap, bypass_short, bypass_error)
                
            if sequence is None:
                st.error(msg)
            else:
                st.success("🎉 Tuyệt vời! Đã xuất lịch thành công dựa trên cấu hình chọn lọc!")
                
                sheet["B2"] = datetime.now().strftime("%Y-%m-%d")
                sheet["A6"] = f"{timedelta_to_str(active_gio)}:00"
                
                thuc_te_secs = target_secs + sai_so
                sheet["D6"] = f"{timedelta_to_str(timedelta(seconds=thuc_te_secs))}:00"
                
                sheet.delete_rows(7, sheet.max_row)
                for i, item in enumerate(sequence):
                    row_idx = 7 + i
                    sheet.cell(row=row_idx, column=3, value=item['name'])
                    sheet.cell(row=row_idx, column=4, value=item['duration_str'])
                    sheet.cell(row=row_idx, column=5, value=item['type'])
                    
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="⬇️ BẤM VÀO ĐÂY TẢI XUỐNG FILE CHÈN HOÀN CHỈNH",
                    data=output,
                    file_name=f"FILE_CHEN_XUAT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                if bypass_short or bypass_error:
                    st.info(f"💡 **Thông tin hệ thống:** File này được tạo ra trong chế độ cưỡng ép.")
                
                if sai_so != 0:
                    trang_thai = "THỪA" if sai_so > 0 else "THIẾU"
                    st.warning(f"⚠️ **LƯU Ý:** Lịch chèn xuất ra thực tế bị **{trang_thai} {abs(sai_so)} giây** so với yêu cầu ban đầu.")
